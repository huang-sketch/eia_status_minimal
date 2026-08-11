import json
import os
import re
from datetime import date, datetime
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Sequence, Tuple

from docx import Document
from openpyxl import load_workbook

from surface_water_pipeline import normalize_factor, normalize_standard_class, parse_numeric_value


NOISE_SHEET_NAME = "噪声监测数据"
SURFACE_WATER_SHEET_NAME = "地表水监测数据"
NOISE_CODE_RE = re.compile(r"^NJ(\d+)(?:-(\d+))?(?:-(\d+))?$", re.IGNORECASE)
EMPTY_MARKERS = {"", "-", "/", "—", "无", "none", "null"}


class InputValidationError(RuntimeError):
    pass


def main() -> None:
    input_dir = Path(os.getenv("EIA_INPUT_DIR", "input"))
    output_dir = Path(os.getenv("EIA_OUTPUT_DIR", "output"))
    run_surface_water = os.getenv("EIA_RUN_SURFACE_WATER", "true").lower() == "true"
    run_noise = os.getenv("EIA_RUN_NOISE", "true").lower() == "true"
    workbook_path = next(
        (path for path in input_dir.glob("*.xlsx") if not path.name.startswith("~$")),
        None,
    )
    plan_path = next(
        (path for path in input_dir.glob("*.docx") if "方案" in path.name and not path.name.startswith("~$")),
        None,
    )
    if workbook_path is None:
        raise FileNotFoundError("未找到 XLSX 监测数据文件")
    if plan_path is None:
        raise FileNotFoundError("未找到 DOCX 监测方案文件")
    parse_xlsx_monitoring_input(
        workbook_path,
        plan_path,
        output_dir,
        run_surface_water=run_surface_water,
        run_noise=run_noise,
    )


def parse_xlsx_monitoring_input(
    workbook_path: Path,
    plan_path: Path,
    output_dir: Path,
    *,
    run_surface_water: bool,
    run_noise: bool,
) -> Dict[str, Any]:
    workbook_path = Path(workbook_path)
    plan_path = Path(plan_path)
    output_dir = Path(output_dir)
    debug_dir = output_dir / "debug_tables"
    extraction_dir = output_dir / "extraction"
    debug_dir.mkdir(parents=True, exist_ok=True)
    extraction_dir.mkdir(parents=True, exist_ok=True)

    workbook = load_workbook(workbook_path, data_only=True, read_only=False)
    plan = parse_monitoring_plan(plan_path)
    errors: List[str] = []
    warnings: List[str] = []
    correspondence: Dict[str, Any] = {}
    summary: Dict[str, Any] = {
        "source_file": str(workbook_path),
        "plan_file": str(plan_path),
        "selected_modules": {
            "surface_water": run_surface_water,
            "noise": run_noise,
        },
        "sheets": workbook.sheetnames,
    }

    if run_noise:
        noise_sheet = find_sheet(workbook, NOISE_SHEET_NAME, ("测点编号", "监测数据"))
        if not plan["noise"]:
            errors.append("监测方案中缺少声环境点位表")
        if noise_sheet is None:
            errors.append(f"缺少工作表: {NOISE_SHEET_NAME}")
            noise_records: List[Dict[str, Any]] = []
        else:
            noise_records, noise_errors, noise_warnings = parse_noise_sheet(noise_sheet)
            errors.extend(noise_errors)
            warnings.extend(noise_warnings)
        code_errors = validate_noise_plan_codes(plan["noise"])
        errors.extend(code_errors)
        noise_match = compare_point_sets(
            "声环境",
            [item["point_code"] for item in plan["noise"]],
            [item["point_code"] for item in noise_records],
        )
        correspondence["noise"] = noise_match
        errors.extend(noise_match["errors"])
        errors.extend(validate_noise_metadata(plan["noise"], noise_records))
        write_noise_flattened(debug_dir / "flattened_table_0.json", noise_records, workbook_path)
        summary["noise_record_count"] = len(noise_records)

    if run_surface_water:
        water_sheet = find_sheet(workbook, SURFACE_WATER_SHEET_NAME, ("测点编号", "水体名称"))
        if not plan["surface_water"]:
            errors.append("监测方案中缺少地表水点位表")
        if water_sheet is None:
            errors.append(f"缺少工作表: {SURFACE_WATER_SHEET_NAME}")
            water_records: List[Dict[str, Any]] = []
            water_points: List[Dict[str, Any]] = []
        else:
            water_records, water_points, water_errors, water_warnings = parse_surface_water_sheet(
                water_sheet,
                workbook_path,
            )
            errors.extend(water_errors)
            warnings.extend(water_warnings)
        water_match = compare_point_sets(
            "地表水",
            [item["point_code"] for item in plan["surface_water"]],
            [item["point_code"] for item in water_points],
        )
        correspondence["surface_water"] = water_match
        errors.extend(water_match["errors"])
        errors.extend(validate_surface_water_metadata(plan["surface_water"], water_points))
        write_json(extraction_dir / "xlsx_surface_water_records.json", water_records)
        summary["surface_water_record_count"] = len(water_records)

    validation = {
        **summary,
        "valid": not errors,
        "errors": unique_strings(errors),
        "warnings": unique_strings(warnings),
    }
    write_json(debug_dir / "point_correspondence.json", correspondence)
    write_json(debug_dir / "xlsx_input_validation.json", validation)
    if errors:
        preview = "；".join(unique_strings(errors)[:8])
        raise InputValidationError(f"XLSX 输入数据校核失败：{preview}")
    return validation


def parse_monitoring_plan(plan_path: Path) -> Dict[str, List[Dict[str, str]]]:
    document = Document(plan_path)
    result: Dict[str, List[Dict[str, str]]] = {"noise": [], "surface_water": []}
    for table in document.tables:
        rows = [[cell.text.strip() for cell in row.cells] for row in table.rows]
        if len(rows) < 2:
            continue
        headers = rows[0]
        if find_header(headers, ("测点编号", "监测点编号")) is None:
            continue
        if find_header(headers, ("测点楼层", "噪声执行标准")) is not None:
            for row in rows[1:]:
                item = {
                    "point_code": normalized_code(value_by_header(headers, row, ("测点编号", "监测点编号"))),
                    "point_name": value_by_header(headers, row, ("测点名称", "监测点名称", "敏感点名称")),
                    "station": value_by_header(headers, row, ("测点桩号", "监测点桩号", "桩号")),
                    "position": value_by_header(headers, row, ("测点位置", "监测点位置", "监测位置")),
                    "floor": value_by_header(headers, row, ("测点楼层", "监测楼层", "楼层")),
                    "duration": value_by_header(headers, row, ("监测时长", "单次监测时间")),
                    "standard_class": value_by_header(headers, row, ("噪声执行标准", "执行标准")),
                }
                if item["point_code"]:
                    result["noise"].append(item)
        elif find_header(headers, ("水体名称", "取样断面", "水质执行标准")) is not None:
            for row in rows[1:]:
                item = {
                    "point_code": normalized_code(value_by_header(headers, row, ("测点编号", "监测点编号"))),
                    "water_name": value_by_header(headers, row, ("水体名称", "河流名称")),
                    "station": value_by_header(headers, row, ("测点位置", "监测点位置", "桩号")),
                    "section": value_by_header(headers, row, ("取样断面", "监测断面")),
                    "standard_class": value_by_header(headers, row, ("水质执行标准", "执行标准", "水质目标")),
                }
                if item["point_code"]:
                    result["surface_water"].append(item)
    return result


def parse_noise_sheet(sheet: Any) -> Tuple[List[Dict[str, Any]], List[str], List[str]]:
    top = [clean_cell(cell.value) for cell in sheet[1]]
    child = [clean_cell(cell.value) for cell in sheet[2]] if sheet.max_row >= 2 else []
    headers = flatten_headers(top, child)
    errors: List[str] = []
    warnings: List[str] = []
    required = {
        "point_code": find_header(top, ("测点编号", "监测点编号", "点位编号")),
        "point_name": find_header(top, ("测点名称", "监测点名称", "敏感点名称")),
        "position": find_header(top, ("测点位置", "监测点位置", "监测位置")),
        "floor": find_header(top, ("测点楼层", "监测楼层", "楼层")),
        "date": find_combined_header(headers, ("监测时间_日期", "监测日期", "日期")),
        "time": find_combined_header(headers, ("监测时间_时间", "测量时间", "时间")),
        "period": find_combined_header(headers, ("监测时间_时段", "时段")),
        "laeq": find_header(top, ("监测数据", "LAeq", "等效声级")),
    }
    missing = [field for field, index in required.items() if index is None]
    if missing:
        return [], ["噪声监测数据缺少关键字段: " + "、".join(missing)], warnings

    station_index = find_header(top, ("测点桩号", "监测点桩号", "桩号"))
    duration_index = find_header(top, ("监测时长", "单次监测时间"))
    standard_indexes = [index for index, value in enumerate(top) if normalized_header(value) == "噪声执行标准"]
    plan_standard_index = standard_indexes[0] if standard_indexes else None
    result_limit_index = standard_indexes[1] if len(standard_indexes) > 1 else None
    vehicle_indexes = {
        label: find_combined_header(headers, (f"车流量数据_{label}", label))
        for label in ("小型车", "中型车", "大型车")
    }

    records: List[Dict[str, Any]] = []
    carry: Dict[int, str] = {}
    seen = set()
    for row_index, row_cells in enumerate(sheet.iter_rows(min_row=3, values_only=True), start=3):
        row = [clean_cell(value) for value in row_cells]
        if not any(row) or str(row[0]).startswith("注意"):
            continue
        for index in range(min(7, len(row))):
            if row[index]:
                carry[index] = row[index]
            elif index in carry:
                row[index] = carry[index]
        point_code = normalized_code(row[required["point_code"]])
        raw_date = row[required["date"]]
        raw_period = row[required["period"]]
        if not point_code and not raw_date:
            continue
        if not NOISE_CODE_RE.fullmatch(point_code):
            errors.append(f"噪声第{row_index}行测点编号不合法: {point_code or '-'}")
            continue
        sample_date = normalize_date(raw_date)
        period = normalize_period(raw_period, row[required["time"]])
        if not sample_date:
            errors.append(f"{point_code} 第{row_index}行监测日期不合法: {raw_date}")
        if period not in {"昼间", "夜间"}:
            errors.append(f"{point_code} 第{row_index}行时段不合法: {raw_period}")
        duplicate_key = (point_code, sample_date, period)
        if duplicate_key in seen:
            errors.append(f"噪声监测数据重复: {point_code} {sample_date} {period}")
        seen.add(duplicate_key)
        laeq, value_warning = parse_numeric_value(row[required["laeq"]])
        if laeq is None:
            errors.append(f"{point_code} {sample_date} {period} 未识别有效噪声值")
        elif value_warning:
            warnings.append(f"{point_code} {sample_date} {period}: {value_warning}")
        vehicle_values = {}
        for vehicle, label in (("large", "\u5927\u578b\u8f66"), ("medium", "\u4e2d\u578b\u8f66"), ("small", "\u5c0f\u578b\u8f66")):
            index = vehicle_indexes.get(label)
            vehicle_values[vehicle] = row[index] if index is not None and index < len(row) else ""
        traffic_flow = (
            "/".join(vehicle_values[vehicle] or "/" for vehicle in ("large", "medium", "small"))
            if any(vehicle_values.values()) else "/"
        )
        records.append(
            {
                "point_code": point_code,
                "point_name": row[required["point_name"]],
                "station": row[station_index] if station_index is not None else "",
                "position": row[required["position"]],
                "floor": row[required["floor"]],
                "duration": row[duration_index] if duration_index is not None else "",
                "standard_class": row[plan_standard_index] if plan_standard_index is not None else "",
                "sample_date": sample_date,
                "time": row[required["time"]],
                "period": period,
                "laeq": laeq,
                "result_limit": parse_optional_number(row[result_limit_index]) if result_limit_index is not None else None,
                "traffic_flow": traffic_flow,
                "traffic_flow_large": vehicle_values["large"] or "/",
                "traffic_flow_medium": vehicle_values["medium"] or "/",
                "traffic_flow_small": vehicle_values["small"] or "/",
                "source_row": row_index,
            }
        )
    errors.extend(validate_noise_period_pairs(records))
    return records, errors, warnings


def parse_surface_water_sheet(
    sheet: Any,
    workbook_path: Path,
) -> Tuple[List[Dict[str, Any]], List[Dict[str, Any]], List[str], List[str]]:
    headers = [clean_cell(cell.value) for cell in sheet[1]]
    errors: List[str] = []
    warnings: List[str] = []
    required = {
        "point_code": find_header(headers, ("测点编号", "监测点编号", "点位编号")),
        "water_name": find_header(headers, ("水体名称", "河流名称")),
        "station": find_header(headers, ("测点位置", "监测点位置", "桩号")),
        "section": find_header(headers, ("取样断面", "监测断面")),
        "date": find_header(headers, ("取样时间", "采样日期", "监测日期")),
        "standard_class": find_header(headers, ("水质执行标准", "执行标准", "水质目标")),
    }
    missing = [field for field, index in required.items() if index is None]
    if missing:
        return [], [], ["地表水监测数据缺少关键字段: " + "、".join(missing)], warnings
    factor_start = max(index for index in required.values() if index is not None) + 1
    records: List[Dict[str, Any]] = []
    points: Dict[str, Dict[str, Any]] = {}
    seen = set()
    for row_index, values in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        row = [clean_cell(value) for value in values]
        if not any(row) or row[0].startswith("注意"):
            continue
        point_code = normalized_code(row[required["point_code"]])
        if not re.fullmatch(r"(?:WJ|D)\d+", point_code, flags=re.IGNORECASE):
            errors.append(f"地表水第{row_index}行测点编号不合法: {point_code or '-'}")
            continue
        sample_date = normalize_date(row[required["date"]])
        if not sample_date:
            errors.append(f"{point_code} 第{row_index}行取样日期不合法: {row[required['date']]}")
        point_meta = {
            "point_code": point_code,
            "water_name": row[required["water_name"]],
            "station": row[required["station"]],
            "section": row[required["section"]],
            "standard_class": row[required["standard_class"]],
        }
        previous = points.get(point_code)
        if previous and normalized_mapping(previous) != normalized_mapping(point_meta):
            errors.append(f"{point_code} 在 XLSX 中存在冲突的点位元数据")
        points[point_code] = point_meta
        for column_index in range(factor_start, len(headers)):
            raw_value = row[column_index] if column_index < len(row) else ""
            if not raw_value:
                continue
            factor, unit = split_factor_header(headers[column_index])
            if not factor:
                continue
            key = (point_code, sample_date, factor)
            if key in seen:
                errors.append(f"地表水监测数据重复: {point_code} {sample_date} {factor}")
                continue
            seen.add(key)
            numeric_value, value_warning = parse_numeric_value(raw_value)
            if numeric_value is None:
                errors.append(f"{point_code} {sample_date} {factor} 未识别有效数值")
            records.append(
                {
                    "source_type": "XLSX监测数据",
                    "monitor_type": "surface_water",
                    "point_code": point_code,
                    "point": point_code,
                    "sample_date": sample_date,
                    "factor": factor,
                    "raw_factor": headers[column_index],
                    "value": raw_value,
                    "numeric_value": numeric_value,
                    "unit": unit,
                    "sample_character": None,
                    "evidence": {headers[index]: row[index] if index < len(row) else "" for index in range(len(headers))},
                    "source_file": str(workbook_path),
                    "source_table": sheet.title,
                    "source_row": row_index,
                    "needs_review": bool(value_warning),
                    "warning": value_warning,
                }
            )
            if value_warning:
                warnings.append(f"{point_code} {sample_date} {factor}: {value_warning}")
    return records, list(points.values()), errors, warnings


def validate_noise_plan_codes(plan_rows: Sequence[Dict[str, str]]) -> List[str]:
    errors: List[str] = []
    grouped: Dict[int, List[Dict[str, str]]] = {}
    for row in plan_rows:
        match = NOISE_CODE_RE.fullmatch(row["point_code"])
        if not match:
            errors.append(f"方案测点编号不符合 NJX-Y-Z 规则: {row['point_code']}")
            continue
        grouped.setdefault(int(match.group(1)), []).append(row)
    name_to_x: Dict[str, int] = {}
    for x, rows in grouped.items():
        names = {normalize_text(row["point_name"]) for row in rows if row["point_name"]}
        if len(names) > 1:
            errors.append(f"NJ{x} 对应了多个测点名称")
        for name in names:
            if name in name_to_x and name_to_x[name] != x:
                errors.append(f"同一测点名称使用了多个敏感点序号: NJ{name_to_x[name]}、NJ{x}")
            name_to_x[name] = x
        expected = expected_noise_codes(x, rows)
        actual = [row["point_code"] for row in rows]
        if actual != expected:
            errors.append(
                f"NJ{x} 编号层级与位置/楼层不一致；期望 {','.join(expected)}，实际 {','.join(actual)}"
            )
    return errors


def expected_noise_codes(x: int, rows: Sequence[Dict[str, str]]) -> List[str]:
    positions = ordered_unique([normalize_text(row["position"]) for row in rows])
    position_order = sorted(
        positions,
        key=lambda value: (position_rank(value), positions.index(value)),
    )
    expected_by_row: Dict[int, str] = {}
    for position_index, position in enumerate(position_order, start=1):
        indexed_rows = [(index, row) for index, row in enumerate(rows) if normalize_text(row["position"]) == position]
        floors = ordered_unique([normalize_floor(row["floor"]) for _, row in indexed_rows])
        meaningful_floors = [floor for floor in floors if floor]
        multiple_floors = len(meaningful_floors) > 1
        floor_order = sorted(meaningful_floors, key=floor_sort_key)
        for row_index, row in indexed_rows:
            if len(position_order) == 1 and not multiple_floors:
                code = f"NJ{x}"
            elif len(position_order) == 1:
                code = f"NJ{x}-{floor_order.index(normalize_floor(row['floor'])) + 1}"
            elif not multiple_floors:
                code = f"NJ{x}-{position_index}"
            else:
                code = f"NJ{x}-{position_index}-{floor_order.index(normalize_floor(row['floor'])) + 1}"
            expected_by_row[row_index] = code
    return [expected_by_row[index] for index in range(len(rows))]


def compare_point_sets(label: str, plan_codes: Iterable[str], data_codes: Iterable[str]) -> Dict[str, Any]:
    plan = sorted(set(filter(None, plan_codes)), key=point_sort_key)
    data = sorted(set(filter(None, data_codes)), key=point_sort_key)
    missing = [code for code in plan if code not in data]
    extra = [code for code in data if code not in plan]
    errors: List[str] = []
    if missing:
        errors.append(f"{label} XLSX 缺少方案点位: " + "、".join(missing))
    if extra:
        errors.append(f"{label} XLSX 存在方案外点位: " + "、".join(extra))
    return {
        "valid": not errors,
        "plan_codes": plan,
        "data_codes": data,
        "missing_in_xlsx": missing,
        "extra_in_xlsx": extra,
        "errors": errors,
    }


def validate_noise_metadata(plan_rows: Sequence[Dict[str, str]], records: Sequence[Dict[str, Any]]) -> List[str]:
    plan_by_code = {row["point_code"]: row for row in plan_rows}
    errors: List[str] = []
    for record in records:
        plan = plan_by_code.get(record["point_code"])
        if not plan:
            continue
        for field, label in (("point_name", "测点名称"), ("station", "桩号"), ("position", "测点位置"), ("floor", "测点楼层")):
            if normalize_text(plan.get(field)) != normalize_text(record.get(field)):
                errors.append(f"{record['point_code']} {label}与方案不一致")
        plan_class = normalize_noise_standard(plan.get("standard_class"))
        data_class = normalize_noise_standard(record.get("standard_class"))
        if plan_class and data_class and plan_class != data_class:
            errors.append(f"{record['point_code']} 噪声执行标准与方案不一致")
        expected_limit = noise_limit(plan_class, record.get("period"))
        result_limit = record.get("result_limit")
        if expected_limit is not None and result_limit is not None and expected_limit != result_limit:
            errors.append(f"{record['point_code']} {record.get('period')}标准限值与方案不一致")
    return unique_strings(errors)


def validate_surface_water_metadata(plan_rows: Sequence[Dict[str, str]], points: Sequence[Dict[str, Any]]) -> List[str]:
    plan_by_code = {row["point_code"]: row for row in plan_rows}
    errors: List[str] = []
    for point in points:
        plan = plan_by_code.get(point["point_code"])
        if not plan:
            continue
        for field, label in (("water_name", "水体名称"), ("station", "测点位置"), ("section", "取样断面")):
            if normalize_text(plan.get(field)) != normalize_text(point.get(field)):
                errors.append(f"{point['point_code']} {label}与方案不一致")
        if normalize_standard_class(plan.get("standard_class")) != normalize_standard_class(point.get("standard_class")):
            errors.append(f"{point['point_code']} 水质执行标准与方案不一致")
    return unique_strings(errors)


def validate_noise_period_pairs(records: Sequence[Dict[str, Any]]) -> List[str]:
    grouped: Dict[Tuple[str, str], set] = {}
    for record in records:
        grouped.setdefault((record["point_code"], record["sample_date"]), set()).add(record["period"])
    errors = []
    for (code, sample_date), periods in grouped.items():
        missing = {"昼间", "夜间"} - periods
        if missing:
            errors.append(f"{code} {sample_date} 缺少" + "、".join(sorted(missing)))
    return errors


def write_noise_flattened(path: Path, records: Sequence[Dict[str, Any]], workbook_path: Path) -> None:
    payload_records = []
    for record in records:
        point = " ".join(
            item
            for item in (
                record["point_code"],
                record.get("point_name"),
                record.get("position"),
                record.get("floor"),
            )
            if item and str(item).lower() not in EMPTY_MARKERS
        )
        monitor_time = " ".join(
            item for item in (record.get("sample_date"), record.get("time"), record.get("period")) if item
        )
        payload_records.append(
            {
                "点位": point,
                "监测时间": monitor_time,
                "区域环境噪声_LAeq": record.get("laeq"),
                "车流量（辆/20min）": record.get("traffic_flow") or "/",
                "\u8f66\u6d41\u91cf\uff08\u8f86/20min\uff09_\u5927\u578b\u8f66": record.get("traffic_flow_large") or "/",
                "\u8f66\u6d41\u91cf\uff08\u8f86/20min\uff09_\u4e2d\u578b\u8f66": record.get("traffic_flow_medium") or "/",
                "\u8f66\u6d41\u91cf\uff08\u8f86/20min\uff09_\u5c0f\u578b\u8f66": record.get("traffic_flow_small") or "/",
                "_source_row_index": record.get("source_row"),
            }
        )
    write_json(
        path,
        {
            "chunk_id": "xlsx_noise_data",
            "source_file": str(workbook_path),
            "table_title": NOISE_SHEET_NAME,
            "header_row_count": 2,
            "header_rows": [],
            "flattened_headers": [
                "点位",
                "监测时间",
                "区域环境噪声_LAeq",
                "车流量（辆/20min）",
                "车流量（辆/20min）_大型车",
                "车流量（辆/20min）_中型车",
                "车流量（辆/20min）_小型车",
            ],
            "data_row_count": len(payload_records),
            "records": payload_records,
        },
    )


def find_sheet(workbook: Any, preferred_name: str, required_headers: Sequence[str]) -> Optional[Any]:
    if preferred_name in workbook.sheetnames:
        return workbook[preferred_name]
    matches = []
    for sheet in workbook.worksheets:
        first_rows = " ".join(clean_cell(cell.value) for row in sheet.iter_rows(min_row=1, max_row=2) for cell in row)
        if all(header in first_rows for header in required_headers):
            matches.append(sheet)
    return matches[0] if len(matches) == 1 else None


def flatten_headers(top: Sequence[str], child: Sequence[str]) -> List[str]:
    headers = []
    for index, parent in enumerate(top):
        part = child[index] if index < len(child) else ""
        if parent and part and normalize_text(parent) != normalize_text(part):
            headers.append(f"{parent}_{part}")
        else:
            headers.append(parent or part)
    return headers


def find_header(headers: Sequence[str], aliases: Sequence[str]) -> Optional[int]:
    normalized = [normalized_header(header) for header in headers]
    alias_values = [normalized_header(alias) for alias in aliases]
    for alias in alias_values:
        if alias in normalized:
            return normalized.index(alias)
    for index, header in enumerate(normalized):
        if any(alias and (alias in header or header in alias) for alias in alias_values):
            return index
    return None


def find_combined_header(headers: Sequence[str], aliases: Sequence[str]) -> Optional[int]:
    return find_header(headers, aliases)


def value_by_header(headers: Sequence[str], row: Sequence[str], aliases: Sequence[str]) -> str:
    index = find_header(headers, aliases)
    return row[index].strip() if index is not None and index < len(row) else ""


def split_factor_header(header: str) -> Tuple[str, str]:
    text = str(header or "").strip()
    match = re.match(r"^(.*?)[（(]\s*([^）)]+)\s*[）)]$", text)
    name = (match.group(1) if match else text).strip()
    unit = (match.group(2) if match else "").strip()
    return normalize_factor(name), unit


def normalize_date(value: Any) -> str:
    if isinstance(value, (date, datetime)):
        return value.strftime("%Y-%m-%d")
    text = clean_cell(value)
    match = re.search(r"(\d{4})[-/.年](\d{1,2})[-/.月](\d{1,2})", text)
    if not match:
        return ""
    try:
        return date(int(match.group(1)), int(match.group(2)), int(match.group(3))).isoformat()
    except ValueError:
        return ""


def normalize_period(value: Any, time_value: Any = "") -> str:
    text = clean_cell(value)
    if "昼" in text:
        return "昼间"
    if "夜" in text:
        return "夜间"
    match = re.search(r"(\d{1,2}):(\d{2})", clean_cell(time_value))
    if match:
        return "昼间" if 6 <= int(match.group(1)) < 22 else "夜间"
    return ""


def normalized_code(value: Any) -> str:
    return re.sub(r"\s+", "", clean_cell(value)).upper()


def normalized_header(value: Any) -> str:
    return re.sub(r"[\s（()）_\-/]+", "", clean_cell(value)).lower()


def normalize_text(value: Any) -> str:
    return re.sub(r"\s+", "", clean_cell(value)).lower()


def normalize_floor(value: Any) -> str:
    text = normalize_text(value)
    return "" if text in EMPTY_MARKERS else text


def normalized_mapping(value: Dict[str, Any]) -> Dict[str, str]:
    return {key: normalize_text(item) for key, item in value.items()}


def clean_cell(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def parse_optional_number(value: Any) -> Optional[float]:
    match = re.search(r"-?\d+(?:\.\d+)?", clean_cell(value))
    return float(match.group(0)) if match else None


def normalize_noise_standard(value: Any) -> str:
    text = normalize_text(value).replace("类区", "类")
    match = re.search(r"(0|1|2|3|4a|4b|4)类?", text, flags=re.IGNORECASE)
    if not match:
        return ""
    token = match.group(1).lower()
    return "4a类" if token == "4" else f"{token}类"


def noise_limit(standard_class: str, period: str) -> Optional[float]:
    day = {"0类": 50, "1类": 55, "2类": 60, "3类": 65, "4a类": 70, "4b类": 70}
    night = {"0类": 40, "1类": 45, "2类": 50, "3类": 55, "4a类": 55, "4b类": 60}
    return (day if period == "昼间" else night if period == "夜间" else {}).get(standard_class)


def position_rank(value: str) -> int:
    for marker, rank in (("首排", 1), ("一排", 1), ("二排", 2), ("三排", 3), ("四排", 4)):
        if marker in value:
            return rank
    return 9999


def floor_sort_key(value: str) -> Tuple[int, str]:
    match = re.search(r"\d+", value)
    if match:
        return (int(match.group(0)), value)
    return (9999, value)


def point_sort_key(value: str) -> Tuple[int, int, int, str]:
    noise = NOISE_CODE_RE.fullmatch(value)
    if noise:
        return (int(noise.group(1)), int(noise.group(2) or 0), int(noise.group(3) or 0), value)
    match = re.search(r"(\d+)$", value)
    return (int(match.group(1)) if match else 9999, 0, 0, value)


def ordered_unique(values: Iterable[str]) -> List[str]:
    result = []
    for value in values:
        if value not in result:
            result.append(value)
    return result


def unique_strings(values: Iterable[str]) -> List[str]:
    return ordered_unique([str(value) for value in values if value])


def write_json(path: Path, payload: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")


if __name__ == "__main__":
    main()
