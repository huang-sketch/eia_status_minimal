import json
import os
import subprocess
import sys
from pathlib import Path

import pytest
from docx import Document
from openpyxl import Workbook

from xlsx_monitoring_parser import (
    InputValidationError,
    parse_xlsx_monitoring_input,
    validate_noise_plan_codes,
)


NOISE_HEADERS = [
    "测点编号",
    "测点名称",
    "测点桩号",
    "测点位置",
    "测点楼层",
    "监测时长",
    "噪声执行标准",
    "监测时间",
    "",
    "",
    "监测数据",
    "噪声执行标准",
    "车流量数据",
    "",
    "",
]
NOISE_CHILD_HEADERS = ["", "", "", "", "", "", "", "日期", "时间", "时段", "", "", "小型车", "中型车", "大型车"]


def add_docx_table(document, headers, rows):
    table = document.add_table(rows=1, cols=len(headers))
    for index, header in enumerate(headers):
        table.cell(0, index).text = header
    for values in rows:
        cells = table.add_row().cells
        for index, value in enumerate(values):
            cells[index].text = str(value)


def build_plan(path: Path, noise_rows):
    document = Document()
    add_docx_table(
        document,
        ["测点编号", "测点名称", "测点桩号", "测点位置", "测点楼层", "监测时长", "噪声执行标准", "备注"],
        noise_rows,
    )
    add_docx_table(
        document,
        ["测点编号", "水体名称", "测点位置", "取样断面", "取样频次", "监测因子", "水质执行标准"],
        [["WJ1", "测试河", "K2+000", "桥梁跨越处", "连续取样两天，每天一次", "水温、pH值、溶解氧", "Ⅲ类"]],
    )
    document.save(path)


def append_noise_measurements(sheet, point):
    code, name, station, position, floor, standard = point
    limits = {"2类": (60, 50), "4a类": (70, 55)}[standard]
    sheet.append([code, name, station, position, floor, "20min", standard, "2026.6.1", "10:00~10:20", "昼间", 58.0, limits[0], 100, 20, 10])
    sheet.append([None, None, None, None, None, None, None, "2026.6.1", "22:00~22:20", "夜间", 48.0, limits[1], 80, 15, 5])


def build_workbook(path: Path, points):
    workbook = Workbook()
    water = workbook.active
    water.title = "地表水监测数据"
    water.append(["测点编号", "水体名称", "测点位置", "取样断面", "取样时间", "水质执行标准", "水温（℃）", "pH值（无量纲）", "溶解氧(mg/L)"])
    water.append(["WJ1", "测试河", "K2+000", "桥梁跨越处", "2026.6.1", "Ⅲ类", 20.0, 7.5, 6.0])
    water.append(["WJ1", "测试河", "K2+000", "桥梁跨越处", "2026.6.2", "Ⅲ类", 21.0, 7.4, 6.2])

    noise = workbook.create_sheet("噪声监测数据")
    noise.append(NOISE_HEADERS)
    noise.append(NOISE_CHILD_HEADERS)
    for point in points:
        append_noise_measurements(noise, point)
    workbook.save(path)


def test_xlsx_pipeline_writes_existing_canonical_inputs(tmp_path):
    plan_path = tmp_path / "监测方案.docx"
    workbook_path = tmp_path / "监测数据.xlsx"
    output_dir = tmp_path / "output"
    noise_rows = [
        ["NJ4-1", "测试小区", "K1+000", "面向项目首排", "2层", "20min", "4a类", ""],
        ["NJ4-2", "测试小区", "K1+000", "面向项目首排", "4层", "20min", "4a类", ""],
    ]
    points = [
        ("NJ4-1", "测试小区", "K1+000", "面向项目首排", "2层", "4a类"),
        ("NJ4-2", "测试小区", "K1+000", "面向项目首排", "4层", "4a类"),
    ]
    build_plan(plan_path, noise_rows)
    build_workbook(workbook_path, points)

    validation = parse_xlsx_monitoring_input(
        workbook_path,
        plan_path,
        output_dir,
        run_surface_water=True,
        run_noise=True,
    )

    assert validation["valid"] is True
    flattened = json.loads((output_dir / "debug_tables" / "flattened_table_0.json").read_text(encoding="utf-8"))
    water_records = json.loads((output_dir / "extraction" / "xlsx_surface_water_records.json").read_text(encoding="utf-8"))
    assert flattened["data_row_count"] == 4
    assert {row["point_code"] for row in water_records} == {"WJ1"}
    assert {row["factor"] for row in water_records} == {"水温", "pH值", "溶解氧"}

    env = os.environ.copy()
    env.update(
        {
            "EIA_INPUT_DIR": str(tmp_path),
            "EIA_OUTPUT_DIR": str(output_dir),
            "EIA_DATA_SOURCE_TYPE": "xlsx_data",
            "ENABLE_LLM_TEXT_POLISH": "false",
            "ENABLE_LLM_EXTRACTION": "false",
            "ENABLE_SCHEMA_FALLBACK": "false",
            "EIA_LLM_DISABLE": "1",
        }
    )
    repository = Path(__file__).resolve().parents[1]
    for script in ("surface_water_pipeline.py", "surface_water_section_generator.py", "noise_section_generator.py"):
        subprocess.run([sys.executable, script], cwd=repository, env=env, check=True, capture_output=True, text=True)
    assert (output_dir / "surface_water_section.docx").exists()
    assert (output_dir / "noise_section.docx").exists()


def test_point_mismatch_is_written_before_generation_is_blocked(tmp_path):
    plan_path = tmp_path / "监测方案.docx"
    workbook_path = tmp_path / "监测数据.xlsx"
    output_dir = tmp_path / "output"
    noise_rows = [
        ["NJ1-1", "测试村", "K1+000", "首排", "2层", "20min", "4a类", ""],
        ["NJ1-2", "测试村", "K1+000", "二排", "2层", "20min", "2类", ""],
    ]
    build_plan(plan_path, noise_rows)
    build_workbook(workbook_path, [("NJ1-1", "测试村", "K1+000", "首排", "2层", "4a类")])

    with pytest.raises(InputValidationError, match="NJ1-2"):
        parse_xlsx_monitoring_input(
            workbook_path,
            plan_path,
            output_dir,
            run_surface_water=False,
            run_noise=True,
        )

    validation = json.loads((output_dir / "debug_tables" / "xlsx_input_validation.json").read_text(encoding="utf-8"))
    correspondence = json.loads((output_dir / "debug_tables" / "point_correspondence.json").read_text(encoding="utf-8"))
    assert validation["valid"] is False
    assert correspondence["noise"]["missing_in_xlsx"] == ["NJ1-2"]


def test_noise_numbering_rules_cover_all_supported_depths():
    rows = [
        {"point_code": "NJ1", "point_name": "单点", "position": "首排", "floor": "2层"},
        {"point_code": "NJ2-1", "point_name": "多位置", "position": "首排", "floor": "2层"},
        {"point_code": "NJ2-2", "point_name": "多位置", "position": "二排", "floor": "2层"},
        {"point_code": "NJ3-1-1", "point_name": "多位置多楼层", "position": "首排", "floor": "2层"},
        {"point_code": "NJ3-1-2", "point_name": "多位置多楼层", "position": "首排", "floor": "4层"},
        {"point_code": "NJ3-2-1", "point_name": "多位置多楼层", "position": "二排", "floor": "2层"},
        {"point_code": "NJ3-2-2", "point_name": "多位置多楼层", "position": "二排", "floor": "4层"},
        {"point_code": "NJ4-1", "point_name": "单位置多楼层", "position": "首排", "floor": "2层"},
        {"point_code": "NJ4-2", "point_name": "单位置多楼层", "position": "首排", "floor": "4层"},
    ]

    assert validate_noise_plan_codes(rows) == []
